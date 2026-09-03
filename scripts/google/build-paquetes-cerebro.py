#!/usr/bin/env python3
"""Build the PAQUETES MODELO 2026 knowledge brain from downloaded Drive files.

Reads raw/ (xlsx/pptx/pdf/png) and writes extracted JSON + CSVs + fichas + informes
under pgt/04-producto/datos/paquetes-modelo-2026/.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

PGT = Path(__file__).resolve().parents[3] / "pgt"
WEB = Path(__file__).resolve().parents[2]
BASE = PGT / "04-producto" / "datos" / "paquetes-modelo-2026"
RAW = BASE / "raw"
OUT_EXT = BASE / "extracted"
OUT_CSV = BASE / "csvs"
OUT_FICHAS = BASE / "fichas"
OUT_INF = BASE / "informes"
TOURS_JSON = WEB / "src" / "content" / "tours"

A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

LANG_FOLDERS = {
    "PROGRAMAS EN INGLES": "en",
    "PROGRAMAS EN ESPAÑOL": "es",
    "PROGRAMAS EN PORTUGES": "pt",
    "Programas con Experiencia": "experiencia",
    "SERVICIOS ADICIONALES": "servicios",
}

# Canonical family keys (duration + english-ish token) → web slug candidates
SLUG_HINTS = {
    "classic-machu-picchu-5d": ["classic machu picchu", "machu picchu classico", "machu picchu clásico", "machu picchu clássico"],
    "machu-picchu-express-3d": ["machu picchu express"],
    "machu-picchu-moderate-4d": ["machu picchu moderate", "machu picchu moderado"],
    "machu-picchu-full-day": ["machu picchu full day", "machu pichu full day"],
    "incredible-machu-picchu-2d": ["incredible machu picchu", "incrível machu picchu", "increible machu picchu"],
    "sacred-valley-machu-picchu-2d": ["valle sagrado + machu picchu", "sacred valley", "valle sagrado, machu picchu"],
    "machu-picchu-humantay-lake-6d": ["humantay", "laguna humantay"],
    "spectacular-cusco-7-days": ["spectacular cusco", "cusco espectacular", "cusco espetacular"],
    "incredible-experience-machu-picchu-7d": ["incredible experience", "increible experiencia", "incrível experiencia"],
    "colonial-lima-and-sacred-land-of-the-incas-7d": ["colonial lima and sacred", "lima colonial y tierra", "lima colonial e terra sagrada"],
    "colonial-lima-huacachina-and-sacred-land-of-the-incas-9-days": ["huacachina"],
    "unforgettable-machu-picchu-8d": ["unforgettable", "inolvidable", "inesquecível", "inesquecivel"],
    "machu-picchu-challenge-8d": ["machu picchu challenge", "desafio machu picchu", "desafío machu picchu"],
    "machu-picchu-extreme-challenge": ["extreme challenge", "desafio extremo", "desafío extremo"],
    "inca-encounters-8d": ["inca encounters", "encuentro de los incas", "encontro dos incas"],
    "short-inca-trail-with-amazon-rainforest-8d": ["short inca trail with amazon", "camino inca corto con la selva"],
    "peru-amazon-rainforest-9d": ["cultura viva", "explore the culture", "eco amazonia", "inkaterra"],
    "gastronomic-and-historic-peru-10d": ["gastronomic", "historico", "histórico y gastron", "histórico e gastron"],
    "origins-of-the-incas-10d": ["origins of the incas", "origen de los", "origens dos incas"],
    "spectacular-peru-10d": ["spectacular peru", "peru espectacular", "peru espetacular"],
    "wonderful-peru-12-days": ["wonderful peru", "peru maravilhoso"],
    "wonder-of-peru-coast-andes-and-rainforest-13d": ["wonder of peru", "maravillas del perú", "costa, andes"],
    "origins-and-mysteries-of-the-andes-16d": ["mysteries of the andes", "mistério de los andes", "mistérios dos andes"],
    "hidden-treasures-huaraz-17d": ["huaraz", "hidden treasures", "tesoros escondidos", "tesouros escondidos"],
    "fascinating-adventure-challenge-in-the-andes-11-days": ["fascinante aventura", "fascinating adventure"],
    "machu-picchu-experience-picnic-with-llamas-7d": ["picnic", "piquenique", "llamas", "lamas"],
    "grand-deluxe-cusco-machu-picchu-by-belmond-5-days": ["by belmond"],
    "grand-deluxe-cusco-machu-picchu-by-casa-andina-hotels-5-days": ["casa andina"],
    "grand-deluxe-cusco-machu-picchu-by-inkaterra-hotels-5-days": ["inkaterra hotels", "by inkaterra"],
    "grand-deluxe-cusco-machu-picchu-by-luxury-collection-hotels-5-days": ["luxury collection"],
    "peru-grand-deluxe-by-belmond-andean-explorer-10-days": ["andean explorer"],
    "peru-grand-deluxe-lima-cusco-machu-picchu-7days": ["peru grand deluxe", "lima, cusco & machu picchu"],
    "short-inca-trail-2d": ["short inca trail", "camino inca corto", "trilha inca curta", "2d camino"],
    "classic-inca-trail-4d": ["classic inca trail", "camino inca clásico", "trilha inca clássica"],
    "classic-inca-trail-7-days": ["camino inca + traslados", "7d camino inca"],
    "sacred-valley-short-inca-trail-3d": ["sacred valley + short", "valle sagrado + camino inca"],
    "salkantay-trek-4-days": ["salkantay sky 4", "4d camino salkantay", "4d trilha salkantay", "4d salkantay"],
    "the-classic-salkantay-trek-5d": ["salkantay sky 5", "5d camino salkantay", "5d trilha salkantay", "5d salkantay"],
    "lares-trek-4d": ["lares"],
    "inca-jungle-trek-4d": ["inca jungle", "jungle trek"],
    "choquequirao-trek-5d": ["choquequirao"],
    "trek-humantay-salkantay-2d": ["humantay lake & salkantay", "lagoa humantay & salkantay", "salkantay pass"],
    "ballestas-huacachina-islands-full-day": ["ballestas", "huacachina"],
    "inti-raymi-full-day": ["inti raymi"],
    "andean-wedding-full-day": ["casamento andino", "casamiento andino", "andean wedding"],
}


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")[:80]


def fold(text: str) -> str:
    text = unicodedata.normalize("NFKD", text or "")
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip().lower()


def detect_lang(rel: str) -> str:
    top = rel.split("/")[0].strip()
    return LANG_FOLDERS.get(top, "other")


def is_copia(name: str) -> bool:
    n = fold(name)
    return n.startswith("copia de") or " - copia" in n or n.endswith("copia")


def duration_from_name(name: str) -> str | None:
    m = re.search(r"\b(\d{1,2})\s*d(?:ias|ías|ays)?\b", fold(name))
    if m:
        return f"{int(m.group(1))}D"
    return None


def guess_slug(folder_name: str, file_name: str) -> str | None:
    blob = fold(f"{folder_name} {file_name}")
    scored: list[tuple[int, str]] = []
    for slug, hints in SLUG_HINTS.items():
        for h in hints:
            if fold(h) in blob:
                scored.append((len(h), slug))
    if not scored:
        return None
    scored.sort(reverse=True)
    return scored[0][1]


def pptx_texts(path: Path, max_slides: int = 80) -> list[dict]:
    slides = []
    try:
        with zipfile.ZipFile(path) as z:
            names = sorted(
                n for n in z.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", n)
            )
            for i, name in enumerate(names[:max_slides], 1):
                xml = z.read(name)
                root = ET.fromstring(xml)
                texts = [t.text or "" for t in root.iter(f"{A_NS}t") if t.text]
                joined = "\n".join(t.strip() for t in texts if t.strip())
                slides.append({"slide": i, "text": joined[:8000]})
    except Exception as e:
        slides.append({"slide": 0, "error": str(e)})
    return slides


def is_day_header(val) -> int | None:
    if isinstance(val, (int, float)) and 1 <= float(val) <= 30 and float(val) == int(val):
        return int(val)
    if isinstance(val, str):
        s = val.strip().replace(",", " ")
        if re.fullmatch(r"\d+(\s+\d+)*", s):
            return int(s.split()[0])
        m = re.match(r"^(\d{1,2})\b", s)
        if m and any(k in fold(s) for k in ("dia", "day", "d ")):
            return int(m.group(1))
    return None


def cell_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def parse_xlsx(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {"sheets": [], "sheet_names": list(wb.sheetnames)}
    for name in wb.sheetnames:
        ws = wb[name]
        days = []
        lines = []
        hotels = []
        current_day = None
        current_title = ""
        nums_f, nums_h, nums_j = [], [], []  # single/double/triple prix x pax-ish
        last_rows = []
        for i, row in enumerate(ws.iter_rows(max_row=200, max_col=16, values_only=True), 1):
            vals = list(row) + [None] * (16 - len(row))
            a, b, c, d = vals[0], vals[1], vals[2], vals[3]
            f, g, h, ii, j, k = vals[5], vals[6], vals[7], vals[8], vals[9], vals[10]
            if all(v is None or str(v).strip() == "" for v in vals[:12]):
                continue
            last_rows.append({"row": i, "vals": [None if v is None else str(v)[:80] for v in vals[:16]]})
            day = is_day_header(a)
            if day is not None and (b is not None) and cell_num(b) is None:
                current_day = day
                current_title = str(b).strip()
                days.append({"day": day, "title": current_title, "row": i})
                continue
            tipo = str(a).strip() if a is not None else ""
            detalle = str(b).strip() if b is not None else ""
            ciudad = str(c).strip() if c is not None else ""
            proveedor = str(d).strip() if d is not None else ""
            if not tipo and not detalle:
                continue
            line = {
                "row": i,
                "day": current_day,
                "tipo": tipo,
                "detalle": detalle,
                "ciudad": ciudad,
                "proveedor": proveedor,
                "single_pax": cell_num(f),
                "single_grp": cell_num(g),
                "double_pax": cell_num(h),
                "double_grp": cell_num(ii),
                "triple_pax": cell_num(j),
                "triple_grp": cell_num(k),
            }
            lines.append(line)
            if "ALOJ" in tipo.upper() or fold(tipo).startswith("htl") or fold(detalle) in {"htl", "hotel", "hotel "}:
                hotels.append(
                    {
                        "day": current_day,
                        "ciudad": ciudad,
                        "hotel": proveedor or detalle,
                    }
                )
            for n in (line["single_pax"],):
                if n:
                    nums_f.append(n)
            for n in (line["double_pax"],):
                if n:
                    nums_h.append(n)
            for n in (line["triple_pax"],):
                if n:
                    nums_j.append(n)

        # FINAL prices: scan last 15 nonempty rows for numeric sequence matching formula labels
        final = {}
        costo_est = {
            "single": round(sum((ln["single_pax"] or 0) + (ln["single_grp"] or 0) for ln in lines), 2),
            "double": round(sum((ln["double_pax"] or 0) + (ln["double_grp"] or 0) for ln in lines), 2),
            "triple": round(sum((ln["triple_pax"] or 0) + (ln["triple_grp"] or 0) for ln in lines), 2),
        }
        # Heuristic from known template: last 4 rows often COSTO / UTILIDAD / BACKUP / FINAL
        labels = []
        for rec in last_rows[-12:]:
            joined = " ".join(v for v in rec["vals"] if v)
            labels.append((rec["row"], joined, rec["vals"]))
        # Try cached FINAL numbers in cols F/H/J near bottom
        for rec in last_rows[-8:]:
            nums = []
            for idx in (5, 7, 9):
                try:
                    nums.append(float(rec["vals"][idx]))
                except (TypeError, ValueError, IndexError):
                    nums.append(None)
            if all(n and n > 50 for n in nums):
                final = {"single": nums[0], "double": nums[1], "triple": nums[2], "source": f"row {rec['row']}"}

        # Recalc using discovered formula: (sum pax + grp/n) * 1.25 * 1.08
        def selling(sum_pax, sum_grp, n_pax, util=0.25, backup=0.08):
            if not sum_pax and not sum_grp:
                return None
            costo = (sum_pax or 0) + ((sum_grp or 0) / n_pax)
            v = costo * (1 + util) * (1 + backup)
            import math

            return int(math.ceil(v - 1e-9)) if v else None

        recalc = {
            "single": selling(
                sum(ln["single_pax"] or 0 for ln in lines),
                sum(ln["single_grp"] or 0 for ln in lines),
                1,
            ),
            "double": selling(
                sum(ln["double_pax"] or 0 for ln in lines),
                sum(ln["double_grp"] or 0 for ln in lines),
                2,
            ),
            "triple": selling(
                sum(ln["triple_pax"] or 0 for ln in lines),
                sum(ln["triple_grp"] or 0 for ln in lines),
                3,
            ),
            "formula": "(sum PRIX_X_PAX + sum PRIX_GRUPO/n) * 1.25 utilidad * 1.08 backup, ROUNDUP",
        }

        out["sheets"].append(
            {
                "name": name,
                "days": days,
                "hotels": hotels,
                "line_count": len(lines),
                "lines": lines,
                "costo_sum_raw": costo_est,
                "precio_final_cached": final,
                "precio_recalc_25_8": recalc,
            }
        )
    wb.close()
    return out


def canonical_key(lang: str, folder: str, duration: str | None) -> str:
    """Language-agnostic key for matching the same product across EN/ES/PT."""
    f = fold(folder)
    f = re.sub(r"^\d{1,2}d\s*", "", f)
    replacements = [
        ("machu picchu", "mp"),
        ("machupicchu", "mp"),
        ("machu pichu", "mp"),
        ("valle sagrado", "sacred-valley"),
        ("sacred valley", "sacred-valley"),
        ("laguna humantay", "humantay"),
        ("humantay lake", "humantay"),
        ("lagoa humantay", "humantay"),
        ("inca trail", "inca-trail"),
        ("camino inca", "inca-trail"),
        ("trilha inca", "inca-trail"),
        ("short inca trail", "inca-trail-short"),
        ("inca corto", "inca-trail-short"),
        ("inca curta", "inca-trail-short"),
        ("salkantay sky", "salkantay"),
        ("classic", "clasico"),
        ("classico", "clasico"),
        ("clásico", "clasico"),
        ("clássico", "clasico"),
        ("spectacular", "espectacular"),
        ("espetacular", "espectacular"),
        ("unforgettable", "inolvidable"),
        ("inesquecivel", "inolvidable"),
        ("inesquecível", "inolvidable"),
        ("incredible experience", "incrivel-exp"),
        ("increible experiencia", "incrivel-exp"),
        ("incrível experiencia", "incrivel-exp"),
        ("origins of the incas", "origen-incas"),
        ("origen de los incas", "origen-incas"),
        ("origens dos incas", "origen-incas"),
        ("mysteries of the andes", "misterios-andes"),
        ("misterio de los andes", "misterios-andes"),
        ("mistérios dos andes", "misterios-andes"),
        ("hidden treasures", "tesoros-huaraz"),
        ("tesoros escondidos", "tesoros-huaraz"),
        ("tesouros escondidos", "tesoros-huaraz"),
        ("colonial lima", "lima-colonial"),
        ("lima colonial", "lima-colonial"),
        ("inca encounters", "encuentro-incas"),
        ("encuentro de los incas", "encuentro-incas"),
        ("encontro dos incas", "encuentro-incas"),
        ("challenge", "desafio"),
        ("desafío", "desafio"),
        ("extreme", "extremo"),
        ("wonderful peru", "peru-maravilhoso"),
        ("peru maravilhoso", "peru-maravilhoso"),
        ("wonder of peru", "maravillas-peru"),
        ("maravillas del peru", "maravillas-peru"),
        ("cultura viva", "cultura-viva"),
        ("explore the culture", "cultura-viva"),
        ("gastronomic", "gastronomico"),
        ("histórico e gastronômico", "gastronomico"),
        ("historico y gastronomico", "gastronomico"),
        ("tours sueltos", "day-tours"),
        ("day tours", "day-tours"),
        ("cusco treks", "treks"),
        ("camino inca", "treks"),
        ("trilhas", "treks"),
        ("peru grand deluxe", "grand-deluxe"),
    ]
    key = f
    for a, b in replacements:
        key = key.replace(a, b)
    key = slugify(key)
    if duration:
        return f"{duration.lower()}-{key}"
    return key


def load_web_tours() -> dict[str, dict]:
    tours = {}
    if not TOURS_JSON.exists():
        return tours
    for p in TOURS_JSON.glob("*.json"):
        data = json.loads(p.read_text(encoding="utf-8"))
        tours[data.get("slug") or p.stem] = {
            "slug": data.get("slug") or p.stem,
            "title": data.get("title"),
            "h1": data.get("h1"),
            "priceFrom": data.get("priceFrom"),
            "duration": data.get("duration"),
            "itinerary_days": len(data.get("itinerary") or []),
            "included": len(data.get("included") or []),
            "excluded": len(data.get("excluded") or []),
            "url": f"/tour/{data.get('slug') or p.stem}/",
        }
    return tours


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main() -> int:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    for d in (OUT_EXT / "xlsx", OUT_EXT / "pptx", OUT_CSV, OUT_FICHAS, OUT_INF):
        d.mkdir(parents=True, exist_ok=True)

    files = [p for p in RAW.rglob("*") if p.is_file() and not p.name.endswith(".partial")]
    print(f"Raw files: {len(files)}", flush=True)
    web = load_web_tours()

    inventario = []
    paquetes_idx = defaultdict(lambda: {"langs": {}, "files": []})
    precios_rows = []
    hotel_rows = []
    itinerary_rows = []
    service_rows = []
    pptx_rows = []
    parse_errors = []

    for path in sorted(files):
        rel = str(path.relative_to(RAW))
        lang = detect_lang(rel)
        parts = Path(rel).parts
        folder = parts[1] if len(parts) > 1 else parts[0]
        # nested deluxe / treks
        if len(parts) > 3:
            folder = "/".join(parts[1:-1])
        elif len(parts) > 2:
            folder = parts[-2]
        name = path.name
        ext = path.suffix.lower()
        copia = is_copia(name)
        duration = duration_from_name(folder) or duration_from_name(name)
        slug = guess_slug(folder, name)
        canon = canonical_key(lang, folder, duration)
        rec = {
            "rel_path": rel,
            "lang": lang,
            "folder": folder,
            "filename": name,
            "ext": ext,
            "bytes": path.stat().st_size,
            "is_copia": "yes" if copia else "no",
            "duration": duration or "",
            "canonical_key": canon,
            "web_slug_guess": slug or "",
            "web_exists": "yes" if slug and slug in web else "no",
        }
        inventario.append(rec)
        paquetes_idx[canon]["langs"].setdefault(lang, {"folder": folder, "files": []})
        paquetes_idx[canon]["langs"][lang]["files"].append(name)
        paquetes_idx[canon]["files"].append(rec)
        if slug:
            paquetes_idx[canon]["web_slug"] = slug

        if ext in {".xlsx", ".xls"} and not copia:
            try:
                parsed = parse_xlsx(path)
                (OUT_EXT / "xlsx" / f"{slugify(lang + '-' + folder) or 'file'}.json").write_text(
                    json.dumps({"source": rel, "lang": lang, **parsed}, ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                # avoid overwrite: include hash
                h = hashlib.md5(rel.encode()).hexdigest()[:8]
                (OUT_EXT / "xlsx" / f"{h}-{slugify(name)}.json").write_text(
                    json.dumps({"source": rel, "lang": lang, "folder": folder, **parsed}, ensure_ascii=False),
                    encoding="utf-8",
                )
                for sh in parsed["sheets"]:
                    cat = sh["name"]
                    cached = sh.get("precio_final_cached") or {}
                    recalc = sh.get("precio_recalc_25_8") or {}
                    precios_rows.append(
                        {
                            "canonical_key": canon,
                            "lang": lang,
                            "folder": folder,
                            "file": name,
                            "categoria": cat,
                            "precio_single_cached": cached.get("single", ""),
                            "precio_double_cached": cached.get("double", ""),
                            "precio_triple_cached": cached.get("triple", ""),
                            "precio_single_recalc": recalc.get("single", ""),
                            "precio_double_recalc": recalc.get("double", ""),
                            "precio_triple_recalc": recalc.get("triple", ""),
                            "days": len(sh.get("days") or []),
                            "lines": sh.get("line_count", 0),
                            "hotels": " | ".join(
                                sorted({f"{h['ciudad']}:{h['hotel']}" for h in sh.get("hotels") or [] if h.get("hotel")})
                            ),
                            "web_slug": slug or "",
                            "web_priceFrom": (web.get(slug) or {}).get("priceFrom", "") if slug else "",
                        }
                    )
                    for d in sh.get("days") or []:
                        itinerary_rows.append(
                            {
                                "canonical_key": canon,
                                "lang": lang,
                                "folder": folder,
                                "categoria": cat,
                                "day": d.get("day"),
                                "title": d.get("title"),
                                "web_slug": slug or "",
                            }
                        )
                    for htl in sh.get("hotels") or []:
                        hotel_rows.append(
                            {
                                "canonical_key": canon,
                                "lang": lang,
                                "folder": folder,
                                "categoria": cat,
                                "day": htl.get("day"),
                                "ciudad": htl.get("ciudad"),
                                "hotel": htl.get("hotel"),
                                "web_slug": slug or "",
                            }
                        )
                    if cat.upper().startswith("TURIST") or cat.upper() in {"TURISTICO", "TURÍSTICO", "TURISTICO."}:
                        for ln in sh.get("lines") or []:
                            service_rows.append(
                                {
                                    "canonical_key": canon,
                                    "lang": lang,
                                    "folder": folder,
                                    "categoria": cat,
                                    "day": ln.get("day"),
                                    "tipo": ln.get("tipo"),
                                    "detalle": ln.get("detalle"),
                                    "ciudad": ln.get("ciudad"),
                                    "proveedor": ln.get("proveedor"),
                                    "double_pax": ln.get("double_pax") or "",
                                    "double_grp": ln.get("double_grp") or "",
                                }
                            )
            except Exception as e:
                parse_errors.append({"file": rel, "error": f"{type(e).__name__}: {e}"})
                print("XLSX FAIL", rel, e, flush=True)

        if ext in {".pptx", ".ppt"} and not copia:
            slides = pptx_texts(path)
            h = hashlib.md5(rel.encode()).hexdigest()[:8]
            (OUT_EXT / "pptx" / f"{h}-{slugify(name)}.json").write_text(
                json.dumps({"source": rel, "lang": lang, "folder": folder, "slides": slides}, ensure_ascii=False),
                encoding="utf-8",
            )
            full = "\n---\n".join(s.get("text", "") for s in slides)
            pptx_rows.append(
                {
                    "canonical_key": canon,
                    "lang": lang,
                    "folder": folder,
                    "file": name,
                    "slides": len(slides),
                    "chars": len(full),
                    "preview": full[:400].replace("\n", " | "),
                    "web_slug": slug or "",
                }
            )

    write_csv(
        OUT_CSV / "inventario-archivos.csv",
        inventario,
        [
            "rel_path",
            "lang",
            "folder",
            "filename",
            "ext",
            "bytes",
            "is_copia",
            "duration",
            "canonical_key",
            "web_slug_guess",
            "web_exists",
        ],
    )
    write_csv(
        OUT_CSV / "precios-por-categoria.csv",
        precios_rows,
        [
            "canonical_key",
            "lang",
            "folder",
            "file",
            "categoria",
            "precio_single_cached",
            "precio_double_cached",
            "precio_triple_cached",
            "precio_single_recalc",
            "precio_double_recalc",
            "precio_triple_recalc",
            "days",
            "lines",
            "hotels",
            "web_slug",
            "web_priceFrom",
        ],
    )
    write_csv(
        OUT_CSV / "itinerarios-dias.csv",
        itinerary_rows,
        ["canonical_key", "lang", "folder", "categoria", "day", "title", "web_slug"],
    )
    write_csv(
        OUT_CSV / "hoteles-por-categoria.csv",
        hotel_rows,
        ["canonical_key", "lang", "folder", "categoria", "day", "ciudad", "hotel", "web_slug"],
    )
    write_csv(
        OUT_CSV / "servicios-turistico.csv",
        service_rows,
        [
            "canonical_key",
            "lang",
            "folder",
            "categoria",
            "day",
            "tipo",
            "detalle",
            "ciudad",
            "proveedor",
            "double_pax",
            "double_grp",
        ],
    )
    write_csv(
        OUT_CSV / "pptx-indice.csv",
        pptx_rows,
        ["canonical_key", "lang", "folder", "file", "slides", "chars", "preview", "web_slug"],
    )

    canon_rows = []
    for key, info in sorted(paquetes_idx.items()):
        langs = ",".join(sorted(info["langs"].keys()))
        slug = info.get("web_slug", "")
        web_t = web.get(slug) if slug else None
        folders = {lg: v["folder"] for lg, v in info["langs"].items()}
        canon_rows.append(
            {
                "canonical_key": key,
                "langs": langs,
                "in_en": "yes" if "en" in info["langs"] else "no",
                "in_es": "yes" if "es" in info["langs"] else "no",
                "in_pt": "yes" if "pt" in info["langs"] else "no",
                "in_experiencia": "yes" if "experiencia" in info["langs"] else "no",
                "folder_en": folders.get("en", ""),
                "folder_es": folders.get("es", ""),
                "folder_pt": folders.get("pt", ""),
                "n_files": len(info["files"]),
                "web_slug": slug,
                "web_title": (web_t or {}).get("title", ""),
                "web_priceFrom": (web_t or {}).get("priceFrom", ""),
                "on_web": "yes" if web_t else "no",
            }
        )
    write_csv(
        OUT_CSV / "paquetes-canonicos.csv",
        canon_rows,
        [
            "canonical_key",
            "langs",
            "in_en",
            "in_es",
            "in_pt",
            "in_experiencia",
            "folder_en",
            "folder_es",
            "folder_pt",
            "n_files",
            "web_slug",
            "web_title",
            "web_priceFrom",
            "on_web",
        ],
    )

    # language coverage matrix
    only_en = [r for r in canon_rows if r["in_en"] == "yes" and r["in_es"] == "no" and r["in_pt"] == "no"]
    only_pt = [r for r in canon_rows if r["in_pt"] == "yes" and r["in_en"] == "no"]
    only_es = [r for r in canon_rows if r["in_es"] == "yes" and r["in_en"] == "no"]
    missing_es = [r for r in canon_rows if r["in_en"] == "yes" and r["in_es"] == "no"]
    missing_pt = [r for r in canon_rows if r["in_en"] == "yes" and r["in_pt"] == "no"]
    missing_web = [r for r in canon_rows if r["web_slug"] and r["on_web"] == "no"]
    drive_not_web = [r for r in canon_rows if not r["web_slug"] or r["on_web"] == "no"]
    web_not_drive = [s for s in web if s not in {r["web_slug"] for r in canon_rows if r["web_slug"]}]

    summary = {
        "generated_at": now,
        "raw_files": len(files),
        "parse_errors": parse_errors,
        "counts": {
            "inventario": len(inventario),
            "canonicos": len(canon_rows),
            "precios_rows": len(precios_rows),
            "itinerary_rows": len(itinerary_rows),
            "hotel_rows": len(hotel_rows),
            "pptx": len(pptx_rows),
            "web_tours": len(web),
        },
        "language_gaps": {
            "only_en": [r["canonical_key"] for r in only_en],
            "only_pt": [r["folder_pt"] or r["canonical_key"] for r in only_pt],
            "only_es": [r["folder_es"] or r["canonical_key"] for r in only_es],
            "en_missing_es": [r["folder_en"] or r["canonical_key"] for r in missing_es],
            "en_missing_pt": [r["folder_en"] or r["canonical_key"] for r in missing_pt],
        },
        "web_gaps": {
            "drive_without_web_slug": [r["canonical_key"] for r in drive_not_web][:80],
            "web_without_drive": web_not_drive,
        },
    }
    (BASE / "RESUMEN.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary["counts"], indent=2))
    print("errors", len(parse_errors))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
